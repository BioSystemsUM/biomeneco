# -*- coding: utf-8 -*-
"""
Created on Mon Jun 29 14:52:26 2020

@author: ecunha
"""
import copy
from os.path import join

import cobra
import os
import warnings
import numpy as np
import pandas as pd
import seaborn
from openpyxl import load_workbook
from sympy import Add

warnings.filterwarnings("ignore")
import re
import matplotlib.pyplot as plt
from cobra import flux_analysis, Model, Reaction, Metabolite
from cobra.flux_analysis import find_essential_genes

class MyModel(Model):
    def __init__(self, file_name = None, biomass_reaction=None, directory=None, prune_mets=False):
        self.pathway_reactions_map = {}
        self.biomass_reaction = None
        if not directory:
            directory = os.getcwd()
        self.directory = directory
        self.file_name = file_name
        self.model_old = []
        self.model_first = None
        self._biomass_composition = None
        self.load_model(self.directory, self.file_name)
        if not biomass_reaction:
            biomass_reaction = self.search_biomass()
        self.bio_reaction = self.model.reactions.get_by_id(biomass_reaction)
        if not self.biomass_reaction:
            self.biomass_reaction = self.bio_reaction.id
        potential_biomet = [e for e in self.bio_reaction.products if 'biomass' in e.id.lower()]
        if potential_biomet:
            self.biomass_metabolite = potential_biomet[0]
        else:
            self.biomass_metabolite = Metabolite("e_Biomass__cytop")
        # self.set_compartments()
        self.reactions_pathways_map = None
        self.bio_precursors = None
        self.pre_precursors = None
        # self.parse_reactions_versions()
        if prune_mets:
            self.model = cobra.manipulation.delete.prune_unused_metabolites(self.model)[0]
        super().__init__(self.model)
        if self.bio_reaction:
            self.objective = self.bio_reaction.id
        self.biomass_components = {}
        # self.parse_genes()
        self.model_first = self.model
        self.get_pathway_reactions_map()
        print("Reactions:", len(self.model.reactions))
        print("Metabolites:", len(self.model.metabolites))
        print("Genes:", len(self.model.genes))
        print("Model loaded")

    @property
    def biomass_composition(self):
        return self._biomass_composition
    @biomass_composition.setter
    def biomass_composition(self, biomass_composition=None):
        if type(biomass_composition) == str or not biomass_composition:
            self._biomass_composition = {}
            if not self.biomass_components:
                self.infer_biomass_from_model()
            biomass_composition = {}
            for children in self.biomass_components[self.biomass_metabolite.id].children:
                biomass_composition[children.id] = children.stoichiometry
            self._biomass_composition = biomass_composition
        elif type(biomass_composition) == Biomass:
            self._biomass_composition = biomass_composition
        else:
            raise Exception("Biomass composition must be a string or a Biomass object")
    def load_model(self, directory, file_name):
        
        ''' This function loads the model. Returns a COBRApy object, the model
        It also creates a copy of the model 
        '''

        print("Loading")
        print("")
        
        os.chdir(directory)
        
        self.model = cobra.io.read_sbml_model(join(directory, file_name))
        # self.model_first = cobra.io.read_sbml_model(os.path.join(directory, file_name))
        if not self.model.exchanges:
            for reaction in self.model.reactions:
                if "EX_" in reaction.id:
                    met = reaction.products[0]
                    reaction.add_metabolites({met:-1})
        # print("deleting b")
        self.delete_b_metabolites()
        return self.model

    def copy(self):
        return self.model.copy()

    def set_compartments(self):
        compartments = self.model.compartments.keys()
        for key in compartments:
            if "outside" in self.model.compartments[key]: self.extra_compartment = key
            if "inside" in self.model.compartments[key]: self.intra_compartment = key
        

    def search_biomass(self):
        for reaction in self.model.reactions:
            if "biomass" in reaction.id.lower():
                return reaction

    def delete_b_metabolites(self):
        for metabolite in self.model.metabolites:
            if '_b' == metabolite.id[-2:] and 'EX_' in metabolite.id:
                self.model.remove_metabolites([metabolite])
                print(metabolite.id + "was removed")

    def get_reaction(self, reaction):
        try:
            return self.reactions.get_by_id(reaction)
        except:
            print(reaction + " not found")

    def get_metabolite(self, metabolite):
        return self.model.metabolites.get_by_id(metabolite)
        
    def get_reactants(self, reaction):
        return self.get_reaction(reaction).reactants
        
    def get_products(self, reaction):
        return self.get_reaction(reaction).products
            
    def get_bio_reac(self):
        return self.bio_reaction
        
    def get_exchanges(self):
        return self.model.exchanges
        
    def get_bio_precursors(self):
        
        ''' This function returns the biomass precursors, i.e. reactants in the biomass reaction '''
        
        self.bio_precursors = []
        reactants = self.get_reactants(self.bio_reaction.id)
        for reactant in reactants:
            self.bio_precursors.append(reactant)
            
        return self.bio_precursors
    
    def get_pre_precursors(self):
        
        ''' This function returns the precursors of the biomass precursors 
        retrieved earlier in the get_bio_precursors() function'''
        try:
            self.pre_precursors = {}
            self.precursors_reactions = {}
            if self.bio_precursors == None: self.get_bio_precursors()
            for precursor in self.bio_precursors:
                reactions = precursor.reactions
                if len(reactions) > 2:
                    print("ATP or other metabolite. Without unique precursor")
                    self.pre_precursors[precursor.id] = []
                else:
                    reaction = ""
                    for r in reactions:
                        if r != self.bio_reaction: reaction = r
                    if reaction:
                        self.pre_precursors[precursor.id] = self.get_reactants(reaction.id)
                        self.precursors_reactions[precursor.name] = (reaction.id, precursor.id)
        except Exception as e:
            print(e)
        return self.pre_precursors
        
    def maximize(self, value = True, pfba = True):
        """ This function maximizes the biomass reaction.
        If value is True, it returns the objective value.
        If value is False, it returns the solution object.
        If pfba is True, it runs pFBA instead of FBA.
        """
        try:
            if value:
                if pfba:
                    return flux_analysis.pfba(self.model)[self.biomass_reaction]
                else:
                    return self.model.optimize().objective_value
            else:
                if pfba:
                    return flux_analysis.pfba(self.model)
                else:
                    return self.model.optimize()
        except Exception as e:
            print(e)
            return 0
    def summary(self, solution=None, pfba = True, **kwargs):
        if solution:
            return self.model.summary(solution, **kwargs)
        if not pfba or kwargs:
            return  self.model.summary(**kwargs)
        else:
            print("Runnning pFBA")
            try:
                pfba_sol = cobra.flux_analysis.pfba(self)
                return self.model.summary(pfba_sol, **kwargs)
            except Exception as e:
                print(e)
                return 0

    def create_reaction(self, reaction):
        self.add_reactions([cobra.Reaction(reaction)])
        return self.get_reaction(reaction)
        
    def create_metabolite(self, metabolite, is_Ex_metabolite = True):
        if is_Ex_metabolite: 
            Compartment = self.extra_compartment
            metabolite_id = "Ex_" + metabolite 
        else: 
            Compartment = self.intra_compartment
            metabolite_id = "In_" + metabolite
        
        self.model.add_metabolites([cobra.Metabolite(metabolite_id, compartment = Compartment)])
        return self.get_metabolite(metabolite_id)
        
    def parse_reactions_versions(self):
        for reaction in self.model.reactions:
            for i in range(10):
                if "_V" + str(i) in reaction.id:
                    try:
                        reaction.id = reaction.id.replace("_V1","")
                        reaction.id = reaction.id.replace("_V2","")
                        reaction.id = reaction.id.replace("_V3","")
                        reaction.id = reaction.id.replace("_V4","")
                        reaction.id = reaction.id.replace("_V5","")
                        reaction.id = reaction.id.replace("_V6","")
                    except:
                        pass
        
        
    def create_exchange(self, metabolite_id, bounds = (-10000,10000)):
        
        
        reaction_name = "Dr_" + metabolite_id
        self.create_reaction(reaction_name).add_metabolites({self.get_metabolite(metabolite_id): -1})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)
        
    def create_transport(self, ex_metabolite, in_metabolite, uptake = True, bounds = (-10000,10000)):
        
        
        if uptake: stoichiometry = 1
        else: stoichiometry = -1
        
        reaction_name = "Tr_" + ex_metabolite + "_" + in_metabolite
        self.create_reaction(reaction_name).add_metabolites({self.get_metabolite(ex_metabolite): -stoichiometry, self.get_metabolite(in_metabolite): stoichiometry})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)
        
    def create_demand(self, metabolite_id):
        
        ''' This function creates a demand reaction'''
        
        reaction_name = "Sk_" + metabolite_id
        self.create_reaction(reaction_name).add_metabolites({self.get_metabolite(metabolite_id): -1})
        self.get_reaction(reaction_name).bounds = (0, 10000)
        return self.get_reaction(reaction_name)
        
    def create_sink(self, metabolite_id, bounds = (-10000, 10000)):
        
        ''' This function creates a sink reaction '''
        
        reaction_name = "Av_" + metabolite_id
        self.create_reaction(reaction_name).add_metabolites({self.get_metabolite(metabolite_id): 1})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)
        
    def create_exchange_transport(self, ex_metabolite, in_metabolite, uptake = True, bounds = (-10000,10000)):
        
        ''' This function creates the transport and exchange reactions at once '''
        
        drain = self.create_exchange(ex_metabolite,bounds)
        transport = self.create_transport(ex_metabolite, in_metabolite, uptake, bounds)
        return drain, transport        
    
    def save(self):
        if len(self.model_old) > 5: self.model_old.pop()
        self.model_old.insert(0,self.model.copy())
        
    def undo(self):
        previous_model = self.model_old.pop(0)
        self.model = previous_model.copy()
        
    def reset(self):
        self.model = self.model_first
    
    def test_bio_precursors(self):
        
        """This function tests any precursor of the biomass. 
        Reactants in the biomass equation"""
        
        if self.bio_precursors == None: self.get_bio_precursors()
        
        self.save()
        
        bio_precursors_res = {"Flux": []}
                              
        meta_val = []
        
        for precursor in self.bio_precursors:
            self.save()
            reaction = self.create_demand(precursor.id)
            self.objective = reaction.id
            bio_precursors_res["Flux"].append(self.maximize())
            meta_val.append(precursor.name)
            self.undo()
            
        self.bio_precursors_res = pd.DataFrame(bio_precursors_res, index=meta_val)
        
        return self.bio_precursors_res
    
    def test_e_precursors(self):
        
        """This function tests any precursor of the biomass (known as e-metabolites), which are
        reactants in the biomass equation.
        Moreover, reactants in the e-metabolite synthesis reaction are also tested
        This function returns a pandas.series data frame"""
            
        if self.bio_precursors == None: self.get_bio_precursors()
        if self.pre_precursors == None: self.get_pre_precursors()
        
        self.save()
        
        e_precursors_res = {"Flux": []}
        meta_val = []
        
        for precursor in self.bio_precursors:
            self.save()
            reaction = self.create_demand(precursor.id)
            self.objective = reaction.id
            val = self.maximize()
            e_precursors_res["Flux"].append(val)
            meta_val.append(precursor.name)
                    
            self.undo()
                    
            for pre_precursor in self.pre_precursors[precursor.id]:
                self.save()
                reaction = self.create_demand(pre_precursor.id)
                self.objective = reaction.id
                val_2 = self.maximize()
                e_precursors_res["Flux"].append(val_2)
                meta_val.append(pre_precursor.name)
            
                self.undo()
            
        self.e_res_precursors = pd.DataFrame(e_precursors_res, index=meta_val)
        
        return self.e_res_precursors
        
    def test_e_reaction(self, e_metabolite):
        
        """The input should be the name of the e_metabolite (e.g. e-Protein_e-Protein)
        This function tests any reactant in the e-metabolite synthesis reaction"""
        
        if self.pre_precursors == None: self.get_pre_precursors()
        
        self.save()
        
        e_precursors_res = {"Flux": []}
        meta_val = []
        
        for pre_precursor in self.pre_precursors[e_metabolite]:
            self.save()
            reaction = self.create_demand(pre_precursor.id)
            self.objective = reaction.id
            val_2 = self.maximize()
            e_precursors_res["Flux"].append(val_2)
            meta_val.append(pre_precursor.name)
        
            self.undo()
        
        res = pd.DataFrame(e_precursors_res, index=meta_val)
        
        return res    
        
    def test_reaction(self, reaction):
        
        """The input should be the id of the reaction (e.g. 00001)
        This function tests any reactant in such reaction"""


        old_objective = self.objective
        precursors = self.get_reactants(reaction)

        # self.save()

        e_precursors_res = {"Flux": []}
        meta_val = []

        for precursor in precursors:
            reaction = self.create_demand(precursor.id)
            self.objective = reaction.id
            val_2 = self.model.optimize().objective_value
            e_precursors_res["Flux"].append(val_2)
            meta_val.append(precursor.name)
            self.remove_reactions(reaction.id)

        res = pd.DataFrame(e_precursors_res, index=meta_val)
        self.objective = old_objective
        return res

    def create_test_drains_to_reaction(self, reaction, reactants = True, products = False, uptake = True, bounds = (-10000,10000)):

            '''This function adds to the model drains or exchange reactions for the reactants or products in a given reaction '''

            self.save()

            if reactants:
                reactant = self.get_reactants(reaction)
                for r in reactant:
                    ex_m = self.create_metabolite(r.id)
                    self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
                print("Drains for all reactants created")

            elif reactants and products:
                reactant = self.get_reactants(reaction)
                product = self.get_products(reaction)
                both = reactant + product
                for r in both:
                    ex_m = self.create_metabolite(r.id)
                    self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
                print("Drains for all reactants and products created")

            else:
                product = self.get_products(reaction)
                for r in product:
                    ex_m = self.create_metabolite(r.id)
                    self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
                print("Drains for all products created")
            
    def create_tRNAs_reactions(self, protein_id="e-Protein"):
        
        ''' This function creates specifically demand reactions for the tRNAs in the aminoacyls forms '''
        
        if self.pre_precursors == None: self.get_pre_precursors()
        
        self.save()

        trnas = self.get_products(self.precursors_reactions[protein_id][0])

        for trna in trnas:
            if "H2O" not in trna.name:
                if "e-Protein" not in trna.name:
                    self.create_sink(trna.id)
                    
        
    
    def create_reporting_file (self, spreadsheet_name):
        
        ''' This function creates an excel spreadsheet 
        with test_e_precursors() output
        '''
        
        res = self.test_e_precursors()
        
        res["Alterations into the model"] = " "
    
        workbook = pd.ExcelWriter(spreadsheet_name)
        res.to_excel(workbook)
        workbook.save()
    
        print(spreadsheet_name, "created")
        
        
    def load_reactions_information(self, file_name):
        
        ''' This function reads the merlin's output file of reactions '''
        
        file = pd.read_excel(file_name)
        
        return file
    
    def create_fluxes (self):
        
        ''' This function run FBA maximizing the objective function and retrieves
        the reactions fluxes'''
        
        self.objective = self.bio_reaction.id
        
        solution = self.optimize()
        
        return solution.fluxes
        
    
    def merge_fluxes (self, reactions_file_name, version):
        
        ''' This function merges merlin's reactions information with the corresponding fluxes '''
        
        fluxes = self.create_fluxes()
        
        merlin_file = self.load_reactions_information(reactions_file_name)
        
        fluxes_ids = list(fluxes.index)

        fluxes_names = [self.get_reaction(reac).name for reac in fluxes_ids]
                        
        fluxes_names_2 = [name.split("__")[0] for name in fluxes_names]
                        
        fluxes_values = [fluxes[val] for val in fluxes_ids]
        
        flux = list(zip(fluxes_ids, fluxes_names_2, fluxes_values))
        
        res = []
        
        res_model_id = []
        
        for f in flux:
            pathway = merlin_file["Pathway Name"][merlin_file["Reaction Name"] == f[1]].values
            equation = merlin_file["Equation"][merlin_file["Reaction Name"] == f[1]].values
            for path in pathway:
                res.append((f[1],path,equation[0],f[2]))
                res_model_id.append(f[0])
        
        df = pd.DataFrame(data = res, index = res_model_id, columns = ["Reaction ID", "Pathway", "Equation", "Flux " + version])
        
        return df
        
    def create_fluxes_spreadsheet(self, reactions_file_name, spreadsheet_name, version):
        
        ''' This function generates an excel spreadsheet with the reactions information and fluxes'''
        
        df = self.merge_fluxes(reactions_file_name, version)
        
        workbook = pd.ExcelWriter(spreadsheet_name)
        df.to_excel(workbook)
        workbook.save()
        
        print(spreadsheet_name, "created")

    def add_medium(self, medium_file_name, medium_sheet_name):
        
        ''' This function constrains the environmental conditions according to the medium composition
        previously defined in a excel spreadsheet file using the exchange reactions
        '''
        
        self.save()
        
        file = pd.read_excel(medium_file_name, medium_sheet_name, converters = {"Model ID": str}, engine = "openpyxl")

        for reaction in file["Reaction ID"]:
            
            reac = self.get_reaction(reaction)
            if reac:
                reac.lower_bound = float(file["LB"][file["Reaction ID"] == reaction].values[0])
                reac.upper_bound = float(file["UB"][file["Reaction ID"] == reaction].values[0])
            else:
                print(reaction, "not found")

    def write(self, filename):
        ''' This function saves the model in a .xml file '''
        cobra.io.write_sbml_model(self, filename)
    
    def update_fluxes_spreadsheet(self, file_name, sheet_name, column_to_write, version, medium):
        
        ''' This function adds new fluxes to an existing excel spreadsheet '''
    
        wb = load_workbook(file_name)
        
        ws = wb[sheet_name]
        
        model_ids = ws["A"]
        
        fluxes = self.create_fluxes()
                        
        for model_id in model_ids:
            
            i = model_id.row
            
            i_2 = column_to_write + str(i)
            
            if i != 1: ws[i_2].value = fluxes[model_id.value]
            else: ws[i_2].value = "Flux " + version + " / " + medium 
        
        wb.save(file_name)
                       
    def get_metabolite_by_name(self, name, compartment = "C00002"):
        found = False
        for met in self.model.metabolites:
            if name == met.name and met.compartment == compartment:
                found = True
                return met
        if not found:
            print(name + "\t"*2+"Not found")
            return None

    def apply_env_conditions_from_excel(self, conditions_file_name, conditions_sheet_name):
        file = pd.read_excel(conditions_file_name, conditions_sheet_name)
        for exchange in self.exchanges:
            try:
                exchange.lower_bound = float(-file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])
            except:
                pass

    def apply_env_conditions_from_dict(self, data, metabolites = None, aliases = None):
            try:
                for metabolite in metabolites:
                    if metabolite in data.keys():
                        if type(data[metabolite]) != tuple:
                            data[metabolite] = (-data[metabolite], 1000)
                        if aliases and metabolite in aliases.keys():
                            metabolite_in_model = aliases[metabolite]
                        else:
                            metabolite_in_model = metabolite
                        self.exchanges.get_by_id(f"EX_{metabolite_in_model}__dra").bounds = data[metabolite]
            except Exception as e:
                print("Error applying environmental conditions")
                print(e)


    def minimal_medium(self, conditions_file_name, conditions_sheet_name, output_file_name, minimal_growth):
        ''' '''

        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        file = pd.read_excel(conditions_file_name, conditions_sheet_name)

        compounds = file["Exchange"].unique()
        exchanges = self.get_exchanges()

        for exchange in exchanges:
            self.get_reaction(exchange.id).lower_bound =  float(-file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])

        exchanges_2 = [exchange for exchange in exchanges]
        output = []
        output_ids = []
        bounds = {}

        for exchange in exchanges:
            print()
            if file["Uptake Rate"][file["Exchange"] == exchange.id].values[0] > 0.01:
                print()
                if exchange.id in compounds:
                    print()
                    bounds[exchange.id] = exchange.lower_bound
                    self.get_reaction(exchange.id).lower_bound = 0.0
                    print()
                    try:
                        biomass_value = self.maximize(pfba=True)
                    except:
                        biomass_value = 0
                        print()

                    self.get_reaction(exchange.id).lower_bound = bounds[exchange.id]

                    if biomass_value <= minimal_growth:
                        output.append((exchange.id, biomass_value, "level 1"))
                        output_ids.append(exchange.id)
                        exchanges_2.remove(exchange)
                else:
                    exchanges_2.remove(exchange)

            else:
                exchanges_2.remove(exchange)
        print()
        for exchange in exchanges_2:
            self.get_reaction(exchange.id).lower_bound = 0.0

        def get_combinations(compounds, exchanges, level, extra_sources, minimal_growth):
            exchange = exchanges[0]
            if exchange.id in compounds:

                self.get_reaction(exchange.id).lower_bound = bounds[exchange.id]

                try:
                    bio_val = self.maximize(pfba=True)
                except:
                    bio_val = 0

                if bio_val > minimal_growth:
                    extra_sources.append((exchange, level + exchange.id))
                    for extra_source in extra_sources:
                        self.get_reaction(extra_source[0].id).lower_bound = 0.0
                    output.append((extra_sources[0][0].name, bio_val, extra_sources[-1][1]))
                    output_ids.append(extra_sources[0][0].id)

                    level = "level 2 ;"
                    extra_sources = []

                else:
                    level = level + exchange.id + " ; "
                    extra_sources.append((exchange, level))
                    ex = exchanges.pop(0)
                    exchanges.append(ex)
                    get_combinations(compounds, exchanges, level, extra_sources, minimal_growth)

        try:
            self.maximize(pfba=True)
            feasible = True
        except:
            feasible = False

        if feasible and self.maximize(pfba=True) > minimal_growth:

            df = pd.DataFrame(data=output, index=output_ids, columns=["Name", "Biomass", "Level"])
            df.to_excel(writer, "Output")
            writer.save()
            writer.close()

        else:
            n_ex = len(exchanges_2)
            exchanges_3 = [exchange for exchange in exchanges_2]

            for i in range(n_ex):
                extra_sources = []
                level = "level 2 ;"
                get_combinations(compounds, exchanges_2, level, extra_sources, minimal_growth)

                reac = exchanges_3.pop(0)
                exchanges_3.append(reac)
                exchanges_2 = [exchange for exchange in exchanges_3]

            df = pd.DataFrame(data=output, index=output_ids, columns=["Name", "Biomass", "Level"])
            df.to_excel(writer, "Output")
            writer.save()
            writer.close()

    def minimize_uptake_sum(self, substrates = None, to_mininimize = None, to_maximize=None):
        if substrates == None:
            substrates = [ex.id for ex in self.exchanges]
        qp_expression = [arg for substrate in substrates for arg in self.get_reaction(substrate).flux_expression.args]

        if to_mininimize:
            qp_expression = [arg for arg in self.get_reaction(to_mininimize).flux_expression.args]
            for arg in qp_expression:
                temp = list(arg.args)
                temp[0] = -1
                arg.args = temp
            qp_expression += [arg for arg in self.get_reaction(to_maximize).flux_expression.args]

        qp_expression = Add(*qp_expression)

        qp_objective = self.problem.Objective(qp_expression, direction='max')

        self.objective = qp_objective


    def gene_essentiality(self, conditions_file_name, conditions_sheet_name):

        file = pd.read_excel(conditions_file_name, conditions_sheet_name)

        for exchange in self.exchanges:
            exchange.lower_bound = float(-file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])

        print(self.maximize(pfba=True))

        deletion_results = find_essential_genes(self, processes=1)

        return deletion_results

    def connectivity(self, output_file_name, extracellular_compartment, cytosol_compartment, periplasm=False):

        # open output file
        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        col = ["Name", "Outside", "Inside"]

        data = {}

        metabolites = self.metabolites

        for met in metabolites:

            idd = str(met.id)

            if met.compartment == extracellular_compartment:
                if idd in data:
                    data[idd][1] = len(met.reactions)
                else:
                    data[idd] = [met.name, len(met.reactions), 0]

            if met.compartment == cytosol_compartment:
                if idd in data:
                    data[idd][2] = len(met.reactions)
                else:
                    data[idd] = [met.name, 0, len(met.reactions)]

        for met in metabolites:
            if periplasm:
                if met.compartment == periplasm:
                    if idd in data:
                        data[idd][1] = data[idd][1] + len(met.reactions)
                    else:
                        data[idd] = [met.name, len(met.reactions), 0]

        df = pd.DataFrame.from_dict(data, orient='index', columns=col)
        df.to_excel(writer, "Output")
        writer.save()
        writer.close()

    def topological_analysis(self, output_file_name):

        # open output file
        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        col = ["Inside", "Outside", "Gene", "Gene Rules"]

        reactions = self.reactions

        outside = 0
        inside = 0
        gene_rules = 0

        for reac in reactions:

            if reac.boundary:
                outside += 1
            else:
                inside += 1
                if reac.gene_reaction_rule != '': gene_rules += 1

        data = {"results": [inside, outside, len(self.genes), gene_rules]}

        df = pd.DataFrame.from_dict(data, orient='index', columns=col)
        df.to_excel(writer, "Output")
        writer.save()
        writer.close()

    def revert_reaction(self, reaction_id):
        reaction = self.get_reaction(reaction_id)
        for met in reaction.metabolites:
            old_st = reaction.metabolites[met]
            reaction.add_metabolites({met: -old_st})
            reaction.add_metabolites({met: -old_st})

    def set_photoautotrophy(self, previous_carbon_source = "EX_C00033__dra", photon_uptake = -855):
        self.exchanges.EX_C00205__dra.bounds = (photon_uptake, 1000)
        self.exchanges.get_by_id(previous_carbon_source).bounds = (0, 1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0, 0)
        self.objective = "e_Biomass__cytop"

    def set_heterotrophy(self, carbon_source = "EX_C00033__dra", update_value = -10):
        self.exchanges.EX_C00205__dra.bounds = (0, 1000)
        self.reactions.get_by_id(carbon_source).bounds = (update_value, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0,1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 0)
        self.objective = "e_Biomass_ht__cytop"

    def set_mixotrophy(self, carbon_source = "EX_C00033__dra", update_value = -10, photon_uptake = -855):
        self.exchanges.get_by_id(carbon_source).bounds = (update_value, 1000)
        self.exchanges.EX_C00205__dra.bounds = (photon_uptake, 1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0, 0)
        self.objective = "e_Biomass__cytop"

    def set_prism_reaction(self, reaction_id):
        for reaction in self.reactions:
            if reaction.id.startswith("PRISM") and reaction.id != reaction_id:
                reaction.bounds = (0, 0)


    def adjust_biomass(self, new_value, suffix="v2", biomass_reaction_id=None):
        if not biomass_reaction_id:
            biomass_reaction_id = self.bio_reaction.id
        new_biomass = Reaction(id=f"e_Biomass_{suffix}__cytop", lower_bound=0, upper_bound=1000)
        if not self.biomass_composition: self.biomass_composition = biomass_reaction_id
        stoichiometries = update_st(copy.deepcopy(self.biomass_composition), new_value)
        new_biomass.add_metabolites({self.metabolites.get_by_id(key): value for key, value in stoichiometries.items()})
        new_biomass.add_metabolites({key: value for key, value in self.reactions.get_by_id(biomass_reaction_id).metabolites.items() if
                                     not key.id.startswith("e_")})
        self.add_reactions([new_biomass])

    def determine_precursors(self, name, composition, units):
        if units == 'mol/mol':
            mol_mol = normalize(composition)
            mg_molMM = {}
            for key, value in mol_mol.items():
                mg_molMM[key] = convert_mmol_mol_to_g_molMM(value, self.metabolites.get_by_id(key).formula_weight)
            mmol_gMM = convert_mg_molMM_to_mmolM_gMM(mol_mol, sum(mg_molMM.values()))
            print(mmol_gMM)

    def adjust_precursors(self, reaction_id, composition, units, suffix="v2"):
        old_composition = self.reactions.get_by_id(reaction_id).metabolites
        for key, value in composition.items():
            if self.metabolites.get_by_id(key) in old_composition:
                old_composition[self.metabolites.get_by_id(key)] = -value / self.metabolites.get_by_id(key).formula_weight
        mol_mol = normalize(old_composition)
        mg_molMM = {}
        for key, value in mol_mol.items():
            mg_molMM[key] = convert_mmol_mol_to_g_molMM(value, key.formula_weight)
        mmol_gMM = convert_mg_molMM_to_mmolM_gMM(mol_mol, sum(mg_molMM.values()))
        new_reaction = Reaction(id=f"{reaction_id.split('__')[0]}_{suffix}__{reaction_id.split('__')[1]}", lower_bound=0, upper_bound=1000)
        for reactant, stoichiometry in mmol_gMM.items():
            new_reaction.add_metabolites({reactant: -round(stoichiometry, 4)})
        new_reaction.add_metabolites({self.metabolites.e_Pigment__chlo: 1})
        self.add_reactions([new_reaction])
        self.reactions.get_by_id(reaction_id).bounds = (0, 0)


    def setup_condition(self, condition):
        for reaction in self.reactions:
            if reaction.id.startswith("e_Biomass") and f"trial{condition}" not in reaction.id:
                reaction.bounds = (0, 0)
            if reaction.id.startswith("e_Pigment") and f"trial{condition}" not in reaction.id:
                reaction.bounds = (0, 0)
        self.objective = f"e_Biomass_trial{condition}__cytop"

    def infer_biomass_from_model(self, biomass_reaction_name="e_Biomass__cytop", biomass_met_id="e_Biomass__cytop"):
        biomass_reaction = self.reactions.get_by_id(biomass_reaction_name)
        macromolecules = biomass_reaction.reactants
        macromolecules = set(macromolecules) - {self.metabolites.get_by_id("C00001__cytop"), self.metabolites.get_by_id("C00002__cytop")}
        if biomass_met_id in [met.id for met in self.metabolites]:
            e_biomass = BiomassComponent(self.metabolites.get_by_id(biomass_met_id), 1, None)
        else:
            e_biomass = BiomassComponent(biomass_met_id, 1, None)
        self.biomass_components[biomass_reaction_name] = e_biomass
        for macromolecule in macromolecules:
            macromolecule_component = BiomassComponent(macromolecule.id, biomass_reaction.metabolites[macromolecule], e_biomass)
            self.biomass_components[macromolecule.id] = macromolecule_component
            get_precursors(macromolecule_component, macromolecule, self)


    def get_reactions_pathways_map(self):
        pathway_map = {}
        for group in self.groups:
            for member in group.members:
                member_id = member.id
                if member_id in pathway_map:
                    pathway_map[member_id].append(group.name)
                else:
                    pathway_map[member_id] = [group.name]
        for reaction in self.reactions:
            if reaction.id not in pathway_map.keys():
                pathway_map[reaction.id] = []
        for key, value in pathway_map.items():
            pathway_map[key] = list(set(value))
        self.reactions_pathways_map = pathway_map

    def get_pathway_reactions_map(self):
        if not self.reactions_pathways_map: self.get_reactions_pathways_map()
        for key, value in self.reactions_pathways_map.items():
            for pathway in value:
                if pathway in self.pathway_reactions_map and key not in self.pathway_reactions_map[pathway]:
                    self.pathway_reactions_map[pathway].append(key)
                else:
                    self.pathway_reactions_map[pathway] = [key]

    def get_genes_pathways_map(self):
        res = {}
        if not self.reactions_pathways_map: self.get_reactions_pathways_map()
        for gene in self.genes:
            reactions = gene.reactions
            reactions_pathways = [self.reactions_pathways_map[reaction.id] for reaction in reactions]
            res[gene.id] = list(set([item for sublist in reactions_pathways for item in sublist]))
        self.genes_pathways_map = res
        return res

    def parse_genes(self):
        to_remove = []
        for gene in self.genes:
            if "_" in gene.id:
                other_gene = self.genes.get_by_id(gene.id.split("_")[0])
                other_gene.annotation = gene.annotation
                if not gene.reactions:
                    to_remove.append(gene)
        cobra.manipulation.remove_genes(self, to_remove)



    def sample(self, method="achr", constraints=None):
        if constraints is None:
            constraints = {}
        for constraint in constraints.items():
            self.reactions.get_by_id(constraint[0]).bounds = constraint[1]

        if method == "achr":
            from cobra.sampling import ACHRSampler
            sampler = ACHRSampler(self, thinning=10, processes=6)
        elif method == "optgp":
            from cobra.sampling import OptGPSampler
            sampler = OptGPSampler(self, thinning=10, processes=6)
        res = [s for s in sampler.batch(100, 10)]
        return res


def test_carbohydrate(model):
    file = open("sugars_to_test.txt")
    sugars=file.readlines()
    file.close()
    for r in model.model.exchanges:
        glucose = model.get_metabolite_by_name("alpha-D-Glucose", "C00001")
        if glucose in r.metabolites:
            r.bounds = (0,0)  
    for sugar in sugars:
        sugar=sugar.replace("\n","")
        if model.get_metabolite_by_name(sugar,"C00001") is not None:
            met = model.get_metabolite_by_name(sugar,"C00001")
            for r in model.model.exchanges:
                if met in r.metabolites:
                    c_number = int(re.search("C(\d*)(.*)",met.formula).group(1))
                    uptake = 6*13/c_number
                    r.bounds = (-uptake,0)
                    biomass = round(model.model.optimize().objective_value,2)
                    r.bounds = (0,0)
            try:
                check_biomass(sugar, biomass)
            except Exception as e:
                print(e)
    

def amino_acid_requirements(model, amino_acids):
    for aa in amino_acids:
        if model.get_metabolite_by_name(aa,"C00001") != None:
            m_laa = model.get_metabolite_by_name(aa,"C00001")
            for r in model.model.exchanges:
                if m_laa in r.metabolites:
                    original_l=r.bounds
                    r.bounds = (0,99999)                  
                    biomass = round(model.model.optimize().objective_value,2)
                    check_biomass(aa, biomass)
                    r.bounds=original_l


def check_biomass(name,biomass):
    if biomass <= 0:
        print(name + "\t"*3 + "No growth")
    elif biomass > 0:
        print(name + "\t"*3 + "Growth" + "\t"*3 + str(biomass))
        
        
def atp_m(model, m_reaction, mu, values=None):
    if values is None:
        values = {0.36: 0, 1: 0, 1.5: 0, 2: 0, 3: 0, 4: 0, 1.48: 0}
    m_reaction = model.get_reaction(m_reaction)
    original_bounds = m_reaction.bounds
    for key in values.keys():    
        m_reaction.bounds = (key,key)
        values[key] = round(model.model.optimize().objective_value,4)
    m_reaction.bounds = original_bounds
    x = np.array(list(values.keys())).reshape((-1, 1))
    y = np.array(list(values.values()))
    from sklearn.linear_model import LinearRegression
    regressor = LinearRegression()  
    regressor.fit(x,y)
    print((mu-regressor.intercept_)/regressor.coef_)
    return values



def get_metabolite_compartment(model):
    c=0
    e=0
    unique_met=[]
    for met in model.model.metabolites:
        if met.compartment == "C_00002":
            c+=1
            unique_met.append(met.name)
        elif met.compartment == "C_00001":
            e+=1
            unique_met.append(met.name)
        
        
    print("Cytoplasmatic:",c,"\nExtracellular:",e,"\nUnique metabolites:", len(set(unique_met)))


def get_transport_number(model):
    tr=[]
    for r in model.model.reactions:
        if r not in model.model.exchanges:
            cyt  = False
            extr = False
            for met in r.metabolites:
                if met.compartment == "e":
                    extr=True
                if met.compartment == "c":
                    cyt = True
            if extr and cyt:
                tr.append(r)
    
    print("Number of transport reactions:",len(set(tr)))


def anaerobic(lr):
    lr.model.reactions.R00209__cytop.bounds = (0,0)
    lr.model.exchanges.EX_C00007__dra.bounds= (0, 99999)
    lr.model.reactions.R00212__cytop.bounds = (0,99999)
    lr.model.reactions.R00258__cytop.bounds = (-99999,0)
    lr.model.reactions.R01827__cytop.bounds = (0,99999)

def aerobic(lr):
    lr.model.exchanges.EX_C00007__dra.bounds= (-3.61, 999999)
    lr.model.reactions.R00209__cytop.bounds = (0,99999)
    lr.model.reactions.R00212__cytop.bounds = (0,0)
    lr.model.reactions.R01827__cytop.bounds = (0,99999)

def define_medium(model,m,sheet="rhamnosus", bigg = False):
    for ex in model.model.exchanges: 
        ex.bounds=(0,99999)
    if bigg:
        data = pd.read_excel("Media_bigg_updated.xlsx",sheet_name = sheet)
    else:
        data = pd.read_excel("Media_backup.xlsx",sheet_name = sheet)
    res = {}
    medium = data["M"+str(m)].dropna().to_list()
    for el in medium:
        res[el]= data["M"+str(m)+"_q"].loc[data["M"+str(m)]==el].values[0]
    for key in res:
        for ex in model.model.exchanges:
            if ex.id == "EX_" + key + "__dra" or ex.id == "EX_" + key + "_e":
                ex.bounds = (-res[key], 99999)
        
    try:
        if m == 3:
            model.model.reactions.R00754__cytop.bounds = (0,0)
    except:
        pass


def aerobic_v2(model):
    o = model.get_metabolite_by_name("Oxygen","extr")
    for ex in model.model.exchanges:
        if o in ex.metabolites:
            ex.bounds = (-99999,99999)
    return model

def growth_carbs(model, m_reaction):
    sugars = ["C00984__extr", "C00243__extr"]
    for r in model.model.exchanges:
        glucose = model.get_metabolite("C00267__cytop")
        if glucose in r.metabolites:
            r.bounds = (0,0)  
    for sugar in sugars:
        for r in model.model.exchanges:
            if sugar in r.metabolites:
                
                if sugar == "C00984__extr":
                    r.bounds = (-2.54,0)
                    print(atp_m(model, m_reaction))
                    r.bounds = (0,0)
                if sugar == "C00243__cytop":
                    r.bounds = (-6.58,0)
                    print(atp_m(model, m_reaction))
                
# growth_carbs(la, atp)



def get_essential(model):
    res=[]
    for r in model.model.reactions:
        original_l=r.bounds
        r.bounds = (0,0)                  
        biomass = round(model.model.optimize().objective_value,2)
        r.bounds=original_l
        if biomass<=0:
            res.append(r)
    return res


def compare_reacts(model1,model2):
    not_in_1=[]
    not_in_2=[]
    reacts_1 = model1.model.reactions
    reacts_2 = model2.model.reactions
    for r in reacts_1:
        if r not in reacts_2:
            not_in_2.append(r.id)
    for r in reacts_2:
        if r not in reacts_1:
            not_in_1.append(r.id)
    file = open("compare_models.csv","w")
    if len(not_in_1) < len(not_in_2):
        for i in range(len(not_in_1)):
            file.write(str(not_in_1[i]) + "\t" + str(not_in_2[i])+ "\n")
        for i in range(len(not_in_1), len(not_in_2)):
            file.write("NAN" + "\t" + str(not_in_2[i])+ "\n")
    if len(not_in_1) >= len(not_in_2):
        for i in range(len(not_in_2)):
            file.write(str(not_in_1[i]) + "\t" + str(not_in_2[i])+ "\n")
        for i in range(len(not_in_2), len(not_in_1)):
            file.write(str(not_in_1[i]) + "\t" + "NAN" + "\n")
    file.close()
    
def define_medium_quercus(model):
    for exchange in model.exchanges:
        exchange.bounds = (0,9999)
    model.exchanges.EX_C00205__dra.bounds = (-100,999)
    model.exchanges.EX_C00011__dra.bounds = (-999,999)
    model.exchanges.EX_C00001__dra.bounds = (-999,999)
    model.exchanges.EX_C00014__dra.bounds = (-999,999)
    model.exchanges.EX_C00080__dra.bounds = (-999,999)
    model.exchanges.EX_C00009__dra.bounds = (-999,999)
    model.exchanges.EX_C00305__dra.bounds = (-999,999)
    model.exchanges.EX_C14818__dra.bounds = (-999,999)
    model.exchanges.EX_C00059__dra.bounds = (-999,999)
    model.exchanges.EX_C00007__dra.bounds = (-999,999)
    

def react_without_gene(model):
    all_reactions = list(model.reactions)
    for r in all_reactions:
        if r in model.exchanges:
            all_reactions.remove(r)
    for gene in model.genes:
        for reaction in gene.reactions:
            if reaction in all_reactions:
                all_reactions.remove(reaction)
    for i in range(len(all_reactions)):
        all_reactions[i] = all_reactions[i].id
    print(all_reactions)
    
def photosynthesis(model):
    copy = model.model.copy()
    copy.reactions.R03140__chlo.bounds = (0,0)
    copy.reactions.R00024__chlo.bounds = (0,9999)
    copy.objective = "EX_C00205__dra"
    return copy

def photorespiration_v1(model,q=3):
    copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
    copy.reactions.R00024__chlo.flux_expression - copy.reactions.R03140__chlo.flux_expression*q,
    lb=0,
    ub=0)
    copy.add_cons_vars(same_flux)
    copy.objective = "EX_C00205__dra"
    copy.reactions.e_Biomass_Leaf__cyto.bounds = (0.1,0.1)
    copy.exchanges.EX_C00205__dra.lower_bound=-1000
    return copy

def photorespiration_v2(model, q=3):
    copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
    copy.reactions.R00024__chlo.flux_expression - copy.reactions.R03140__chlo.flux_expression*q,
    lb=0,
    ub=0)
    copy.add_cons_vars(same_flux)
    copy.objective = "e_Biomass_Leaf__cyto"
    copy.exchanges.EX_C00205__dra.bounds = (-100,999)
    return copy


def update_biomass(model, biomass_reaction, metabolites_to_remove):
    """
    NOT FINISHED
    """
    for met_id in metabolites_to_remove:
        metabolite = model.metabolites.get_by_id(met_id)
        if metabolite in model.reactions.get_by_id(biomass_reaction).metabolites:
            st = model.reactions.get_by_id(biomass_reaction).metabolites[metabolite]
            model.reactions.get_by_id(biomass_reaction).add_metabolites({model.metabolites.get_by_id(metabolite): -st})
    for reactant in  model.reactions.get_by_id(biomass_reaction).reactants:
        pass



def photorespiration_mt(model,q):
    copy = model.copy()
    same_flux = copy.problem.Constraint(
        copy.reactions.R00024__plst_Leaf_Light.flux_expression - copy.reactions.R03140__plst_Leaf_Light.flux_expression * q,
        lb=0,
        ub=0)

    same_flux1 = copy.problem.Constraint(
        copy.reactions.R00024__plst_Leaf_Dark.flux_expression - copy.reactions.R03140__plst_Leaf_Dark.flux_expression * q,
        lb=0,
        ub=0)

    same_flux2 = copy.problem.Constraint(
        copy.reactions.R00024__plst_Phellogen_Light.flux_expression - copy.reactions.R03140__plst_Phellogen_Light.flux_expression * q,
        lb=0,
        ub=0)

    same_flux3 = copy.problem.Constraint(
        copy.reactions.R00024__plst_Phellogen_Dark.flux_expression - copy.reactions.R03140__plst_Phellogen_Dark.flux_expression * q,
        lb=0,
        ub=0)

    same_flux4 = copy.problem.Constraint(
        copy.reactions.R00024__plst_Ibark_Light.flux_expression - copy.reactions.R03140__plst_Ibark_Light.flux_expression * q,
        lb=0,
        ub=0)
    same_flux5 = copy.problem.Constraint(
        copy.reactions.R00024__plst_Ibark_Dark.flux_expression - copy.reactions.R03140__plst_Ibark_Dark.flux_expression * q,
        lb=0,
        ub=0)

    copy.add_cons_vars(same_flux)
    copy.add_cons_vars(same_flux1)
    copy.add_cons_vars(same_flux2)
    copy.add_cons_vars(same_flux3)
    copy.add_cons_vars(same_flux4)
    copy.add_cons_vars(same_flux5)
    same_flux6 = copy.problem.Constraint(
        copy.reactions.EX_C00244__dra_Light.flux_expression *2 - copy.reactions.EX_C00244__dra_Dark.flux_expression * 3,
        lb=0,
        ub=0)
    copy.add_cons_vars(same_flux6)

    # same_flux7 = copy.problem.Constraint(
    #     copy.reactions.T_Sucrose__cp1_Leaf_Light.flux_expression * 1 - copy.reactions.T_Sucrose__cp1_Leaf_Dark.flux_expression * 3,
    #     lb=0,
    #     ub=0)
    # copy.add_cons_vars(same_flux7)
    copy.objective = "EX_C00205__dra_Light"
    copy.reactions.EX_C00205__dra_Light.bounds = (-1000,1000)
    copy.reactions.Total_biomass.bounds = (0.11,0.11)
    return copy


def photorespiration_dn(model, q=3):
    copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
    copy.reactions.R00024__chlo_Light.flux_expression - copy.reactions.R03140__chlo_Light.flux_expression*q,
    lb=0,
    ub=0)
    same_flux2 = model.model.problem.Constraint(
        copy.reactions.R00024__chlo_Dark.flux_expression - copy.reactions.R03140__chlo_Dark.flux_expression * q,
        lb=0,
        ub=0)
    copy.add_cons_vars(same_flux)
    copy.add_cons_vars(same_flux2)
    # copy.objective = "e_Biomass_Leaf__cyto"
    # copy.exchanges.EX_C00205__dra.bounds = (-100,999)
    return copy


def photorespiration_v3(model,q=3):
    copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
    copy.reactions.R00024__chlo.flux_expression - copy.reactions.R03140__chlo.flux_expression*q,
    lb=0,
    ub=0)
    copy.add_cons_vars(same_flux)
    copy.objective = "e_Biomass__cyto"
    copy.exchanges.EX_C00205__dra.bounds = (-100,0)
    return copy



def respiration(model):
    copy = model.model.copy()
    copy.exchanges.EX_C00205__dra.bounds=(0,999)
    copy.exchanges.EX_C00089__dra.bounds=(-9999,0)
    copy.objective = "EX_C00089__dra"
    return copy


def get_dataframe(model):
    res = pd.DataFrame()
    res2 = pd.DataFrame()
    res3 = pd.DataFrame()
    res4 = pd.DataFrame()
    res5 = pd.DataFrame()
    res6 = pd.DataFrame()
    calvin_cycle = ["R00024", "R01512","R01063","R01015_V2", "R01070","R04780", "R01830", "R01829",
                     "R01845","R01641", "R01056", "R01523", "R01529"]
    photorespirationp = ["R00372__pero", "R00475__pero", "R00588__pero", "R01221__mito",
                       "R01334__chlo", "R01388__pero", "R01514__chlo", "R03140__chlo",
                        "R00009__pero"]
    TCA = ["R00342__mito", "Cytochrome_C_Oxidase__mito", "ATP_Synthase__mito", 
           "Cytochrome_C_Reductase__mito", "NADH_Dehydrogenase__mito"]
    GS_GOGAT = ["Ferredoxin_NADP_Reductase__chlo", "R00021__chlo",
                "R00253__chlo"]
    mva = ["R00238__cyto", "R01978__cyto", "R02082__cyto", "R02245__cyto", "R03245__pero", "R01121__pero"]
    mep = ["R05636__chlo", "R05688__chlo", "R05633__chlo", "R05634__chlo", "R05637__chlo", "R08689__chlo"]
    for i in range(len(calvin_cycle)):
        calvin_cycle[i] += "__chlo"        
    for i in range(1,6):
        pr = photorespiration_v2(model, i)
        fba = flux_analysis.pfba(pr)
        temp = fba.fluxes.loc[fba.fluxes.index.isin(calvin_cycle)]
        temp2 = fba.fluxes.loc[fba.fluxes.index.isin(photorespirationp)]
        temp3 = fba.fluxes.loc[fba.fluxes.index.isin(TCA)]
        temp4 = fba.fluxes.loc[fba.fluxes.index.isin(GS_GOGAT)]
        temp5 = fba.fluxes.loc[fba.fluxes.index.isin(mva)]
        temp6 = fba.fluxes.loc[fba.fluxes.index.isin(mep)]
        res = pd.concat([res,temp], axis = 1)
        res2 = pd.concat([res2,temp2], axis = 1)
        res3 = pd.concat([res3,temp3], axis = 1)
        res4 = pd.concat([res4,temp4], axis = 1)
        res5 = pd.concat([res5,temp5], axis = 1)
        res6 = pd.concat([res6,temp6], axis = 1)
        res.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res2.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res3.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res4.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res5.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res6.rename({"fluxes": str(i)}, inplace = True, axis=1)
    final_res = pd.concat([res,res2,res3,res4, res5, res6])
    return final_res


def get_biomass_co2_plot(model):
    res_co2 = {}
    res_biomass= {}
    for i in range(1,11):
        pr = photorespiration_v2(model, i)
        fba = pr.optimize()
        res_co2[i] = fba["EX_C00011__dra"]

def check_under_limit(reaction):
    balance= reaction.check_mass_balance()
    if balance:
        for key in balance:
            if round(balance[key],5) != 0:
                return False
    return True


def check_balance(model, show_biomass_reactions=False):
    res = {}
    for reaction in model.reactions:
        if reaction.check_mass_balance() and "EX_" not in reaction.id and not check_under_limit(reaction) and "DM_" not in reaction.id:
            if str(reaction.id.split('__')[0]) + str(reaction.check_mass_balance()) not in res:
                if str(reaction.id).startswith("e_"):
                    if show_biomass_reactions:
                        res[str(reaction.id)] =  reaction.check_mass_balance()
                else:
                    res[str(reaction.id)] = reaction.check_mass_balance()
    return res

def simulation_for_conditions(model, conditions_df, growth_rate_df, save_in_file=False, filename=None, objective=None):
    as_dict = conditions_df.to_dict(orient='index')
    growth_rate = growth_rate_df.to_dict(orient='index')
    complete_results = {}
    error_sum = 0
    values_for_plot = {}
    model.exchanges.EX_C00011__dra.bounds = (-1000, 1000)
    t = model.optimize()
    for index, condition in as_dict.items():
        copy = model.copy()
        for reaction in copy.reactions:
            if "Biomass" in reaction.id and "EX_" not in reaction.id and reaction.id != f"e_Biomass_trial{index}__cytop":
                reaction.bounds = (0, 0)
        copy.reactions.get_by_id(f"e_Biomass_trial{index}__cytop").bounds = (0, 1000)
        if objective:
            [setattr(x, 'objective_coefficient', 0) for x in model.reactions if x.objective_coefficient != 0]
            copy.reactions.get_by_id(f"e_Biomass_trial{index}__cytop").objective_coefficient = 1
            for key, value in objective.items():
                copy.reactions.get_by_id(key).objective_coefficient = value
        else:
            copy.objective = f"e_Biomass_trial{index}__cytop"
        t = copy.optimize()
        for met, lb in condition.items():
            lb = -lb if lb < 0 else lb
            copy.reactions.get_by_id("EX_" + met + "__dra").bounds = (round(-lb, 4), 1000)
        sol = copy.optimize()
        biomass = round(sol[f"e_Biomass_trial{index}__cytop"], 3)
        error_sum += abs(growth_rate[index]['growth_rate'] - biomass)
        complete_results[index] = sol
        values_for_plot[index] = (growth_rate[index]['growth_rate'], biomass)
    if save_in_file:
        write_simulation(complete_results, filename)
    return complete_results, values_for_plot, round(error_sum, 6)





def get_reactions_nadh_nadph(model):
    print("starting.....")
    for reaction in model.reactions:
        if model.metabolites.C00003__cytop in reaction.metabolites:
            for reaction2 in model.reactions:
                if reaction.genes == reaction2.genes and model.metabolites.C00005__cytop in reaction2.metabolites:
                    mets1, mets2 = [], []
                    for met in reaction.metabolites:
                        if met.id != "C00003__cytop" and met.id != "C00004__cytop":
                            mets1.append(met.id)
                    for met in reaction2.metabolites:
                        if met.id != "C00005__cytop" and met.id != "C00006__cytop":
                            mets2.append(met.id)
                    if mets1 == mets2:
                        print(reaction.id, reaction.name)
                        print(reaction2.id, reaction.name)
                        print("----------")


def add_reaction_string_to_dataframe(dataframe, model):
    dataframe['Reaction'] = np.nan
    for reaction in dataframe.index:
        reac = model.reactions.get_by_id(reaction)
        reac_string = ''
        for reactant in reac.reactants:
            reac_string += str(abs(reac.get_coefficient(reactant.id))) + ' ' + reactant.name + ' + '
        reac_string = reac_string[0:len(reac_string)-3]
        if reac.reversibility:
            reac_string += ' <=> '
        else:
            reac_string += ' => '
        for products in reac.products:
            reac_string += str(abs(reac.get_coefficient(products.id))) + ' ' + products.name + ' + '
        if reac_string.strip()[-1] == '+':
            reac_string = reac_string[0:len(reac_string)-3]

        dataframe['Reaction'].loc[dataframe.index == reaction] = reac_string
    return dataframe



def get_heatmap(model):
    res = get_dataframe(model)
    ax = seaborn.clustermap(res, cmap ="YlGnBu", col_cluster = False, row_cluster = False, z_score = 0)
    
    
def count_reactions_by_compartment(model):
    compartments = {}
    for reaction in model.model.reactions:
        
        if len(reaction.compartments)>1:
            membrane = check_transport(reaction)
            if membrane is not None:
                if membrane not in compartments.keys():
                    compartments[membrane] = 1
                else:
                    compartments[membrane] += 1
        else:
            reaction_compartment = list(reaction.compartments)[0]
            if str(reaction.id).endswith('extr_b') is False: 
                compartment = model.model.compartments[reaction_compartment]
                if compartment not in compartments.keys():
                    compartments[compartment] = 1
                else:
                    compartments[compartment] += 1
            
    return compartments

 

def check_transport(reaction):
    d = {"['C_00004', 'C_00006']": 'Chloroplast_Membrane', "['C_00001', 'C_00004']": 'Plasma_Membrane', 
         "['C_00002', 'C_00004']": 'Mitochondrial_Membrane', "['C_00003', 'C_00004']": 'Peroxisome_Membrane',
         "['C_00004', 'C_00008']": 'Vacuole_Membrane', "['C_00004', 'C_00005']": 'Golgi_Membrane', 
         "['C_00004', 'C_00007']": 'Endoplasmic_Reticulum_Membrane', 
         "['C_00006', 'C_00007']": 'Endoplasmic_Reticulum_Membrane'}
    
    compartments = list(reaction.compartments)
    compartments.sort()
    if str(compartments) not in d.keys():
        print(reaction)
        return
    else:
        membrane = d[str(compartments)]
        return membrane
    
    
def get_dataframe_all(model):
    res = pd.DataFrame()   
    for i in range(1,6):
        pr = photorespiration_v2(model, i)
        fba = pr.optimize()
        res = pd.concat([res,fba.fluxes], axis = 1)
        res.rename({"fluxes": str(i)}, inplace = True, axis=1)
    return res


def get_dark_reactions(reactions):
    for i in range(len(reactions)):
        reactions[i] = reactions[i] + "_Leaf_Light"
    return reactions


def get_dataframe_3(model):
    res = pd.DataFrame()
    res2 = pd.DataFrame()
    res3 = pd.DataFrame()
    res4 = pd.DataFrame()
    calvin_cycle = ["R00024", "R01512","R01063","R01015","R04780", "R01830", "R01829",
                     "R01845","R01641", "R01056", "R01523", "R01529"]
    photorespiration = ["R00372__pero", "R00475__pero", "R00588__pero",
                       "R01334__chlo", "R01388__pero", "R01514__chlo", "R03140__chlo",
                        "R00009__pero"]
    TCA = ["Cytochrome_C_Oxidase__mito", "ATP_Synthase__mito",
           "Cytochrome_C_Reductase__mito", "NADH_Dehydrogenase__mito"]
    GS_GOGAT = ["Ferredoxin_NADP_Reductase__chlo", "R00021__chlo",
                "R00253__chlo"]
    for i in range(len(calvin_cycle)):
        calvin_cycle[i] += "__chlo"
    calvin_cycle = get_dark_reactions(calvin_cycle)
    photorespiration = get_dark_reactions(photorespiration)
    TCA = get_dark_reactions(TCA)
    GS_GOGAT = get_dark_reactions(GS_GOGAT)

    for i in range(1,6):
        pr = photorespiration_mt(model, i)
        pfba = flux_analysis.pfba(pr)
        temp = pfba.fluxes.loc[pfba.fluxes.index.isin(calvin_cycle)]
        temp2 = pfba.fluxes.loc[pfba.fluxes.index.isin(photorespiration)]
        temp3 = pfba.fluxes.loc[pfba.fluxes.index.isin(TCA)]
        temp4 = pfba.fluxes.loc[pfba.fluxes.index.isin(GS_GOGAT)]
        res = pd.concat([res,temp], axis = 1)
        res2 = pd.concat([res2,temp2], axis = 1)
        res3 = pd.concat([res3,temp3], axis = 1)
        res4 = pd.concat([res4,temp4], axis = 1)
        res.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res2.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res3.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res4.rename({"fluxes": str(i)}, inplace = True, axis=1)
    final_res= change_name([res,res2,res3,res4])
    return final_res


def get_dataframe_2(model):
    res = pd.DataFrame()
    res2 = pd.DataFrame()
    res3 = pd.DataFrame()
    res4 = pd.DataFrame()
    res5 = pd.DataFrame()
    calvin_cycle = ["R00024","R01641", "R01512","R01063", "R01830", "R01829", #"R01070",  "R04780","R01015","R01056","R01845",
                       "R01523", "R01529"]
    photorespiration = ["R00372__pero", "R00475__pero", "R00588__pero", "R01221__mito",
                       "R01334__chlo", "R01388__pero", "R01514__chlo", "R03140__chlo",
                        "R00009__pero"]
    TCA = ["R00342__mito","R00351__mito", "Cytochrome_C_Oxidase__mito", "ATP_Synthase__mito",
           "Cytochrome_C_Reductase__mito", "NADH_Dehydrogenase__mito"]
    GS_GOGAT = ["Ferredoxin_NADP_Reductase__chlo", "R00021__chlo",
                "R00253__chlo"]
    mva = ["R00238__cyto", "R01978__cyto", "R02082__cyto", "R02245__cyto", "R03245__pero", "R01121__pero"]
    mep = ["R05636__chlo", "R05688__chlo", "R05633__chlo", "R08689__chlo"] # "R05634__chlo", "R05637__chlo",
    for i in range(len(calvin_cycle)):
        calvin_cycle[i] += "__chlo"        
    for i in range(1,6):
        pr = photorespiration_v2(model, i)
        pfba = flux_analysis.pfba(pr)
        temp = pfba.fluxes.loc[pfba.fluxes.index.isin(calvin_cycle)]
        temp2 = pfba.fluxes.loc[pfba.fluxes.index.isin(photorespiration)]
        temp3 = pfba.fluxes.loc[pfba.fluxes.index.isin(TCA)]
        temp4 = pfba.fluxes.loc[pfba.fluxes.index.isin(GS_GOGAT)]
        temp5 = pfba.fluxes.loc[pfba.fluxes.index.isin(mep)]
        res = pd.concat([res,temp], axis = 1)
        res2 = pd.concat([res2,temp2], axis = 1)
        res3 = pd.concat([res3,temp3], axis = 1)
        res4 = pd.concat([res4,temp4], axis = 1)
        res5 = pd.concat([res5,temp5], axis = 1)
        res.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res2.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res3.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res4.rename({"fluxes": str(i)}, inplace = True, axis=1)
        res5.rename({"fluxes":str(i)}, inplace =True, axis =1)
    final_res= change_name([res,res2,res3,res4, res5])
    return final_res


def get_pfba(model):
    try:
        fba = flux_analysis.pfba(model)
        if fba:
            return fba
        else:
            get_pfba(model)
    except:
        fba = flux_analysis.pfba(model)
        return fba

def get_inner_bark_model(model, conditions_file_name, conditions_sheet_name):
    model.apply_env_conditions_from_excel(conditions_file_name, conditions_sheet_name)
    biomass_reactions = ["e_Biomass_Leaf__cyto", "e_Carbohydrate__cyto", "e_CellWall_Leaf__cyto", "e_Cofactor_Leaf__cyto"]
    model.model.objective = 'e_Biomass_Ibark__cyto'
    try:
        for reaction in biomass_reactions:
            model.model.reactions.get_by_id(reaction).bounds = (0,0)
    except Exception as e:
        print(e)


def get_phellogen_model(model, conditions_file_name, conditions_sheet_name):
    model.apply_env_conditions_from_excel(conditions_file_name, conditions_sheet_name)
    biomass_reactions = ["e_Biomass_Leaf__cyto", "e_Carbohydrate_Leaf__cyto", "e_CellWall_Leaf__cyto", "e_Cofactor_Leaf__cyto",
                         "e_Biomass_Ibark__cyto", "e_Carbohydrate_Ibark__cyto", "e_CellWall_Ibark__cyto", "e_Cofactor_Ibark__cyto","e_Suberin_Ibark__cyto"]
    model.model.objective = 'e_Biomass_Phellogen__cyto'
    try:
        for reaction in biomass_reactions:
            model.model.reactions.get_by_id(reaction).bounds = (0,0)
    except Exception as e:
        print(e)
    

def change_name(dfs):
    for df in dfs:
        index = df.index.values
        for i in index:
            df.rename(index = {i: i.replace("ATP_Synthase", "ATPS").replace("Cytochrome_C_Reductase","CCOR").replace("Cytochrome_C_Oxidase","COX").replace("NADH_Dehydrogenase","NAD9").replace("Ferredoxin_NADP_Reductase","FNR").replace("__chlo","").replace("__pero","").replace("__mito","")}
                              ,inplace=True)
    return dfs

def build_heatmap(model):
    res,res2,res3,res4,res5 = get_dataframe_2(model)
    normalized1=res.div(res["3"],axis=0)
    normalized2=res2.div(res2["3"],axis=0)
    normalized3=res3.div(res3["3"],axis=0)
    normalized4=res4.div(res4["3"],axis=0)
    normalized5 = res5.div(res5["3"], axis=0)
    vmin = 9999
    vmax = -9999
    res = [normalized1,normalized2,normalized3,normalized4,normalized5 ]
    for df in res:
        temp_min = df.min().min()
        temp_max = df.max().max()
        if temp_min < vmin:
            vmin=temp_min
        if temp_max> vmax:
            vmax=temp_max
    grid_kws = {"height_ratios": (5,5,5,5,1), "hspace": .3}
    f, (ax, ax2,ax3,ax5,cbar_ax) = plt.subplots(5,gridspec_kw=grid_kws,figsize = (7,18))
    ax.tick_params(colors='k', grid_color='k')
    f.tight_layout(pad=0.0001)
    ax1 = seaborn.heatmap(normalized1, ax=ax, cmap ="YlGnBu",
                 cbar=False, vmin = vmin, vmax = vmax,
                 cbar_kws={"orientation": "horizontal"}) 
    ax1.set_ylabel("Calvin Cycle", rotation = 0, fontsize = 16, color="k")
    ax1.yaxis.set_label_coords(-0.35, 0.4)
    ax1.axes.xaxis.set_visible(False)
    pos1 = ax1.get_position()
    pos = [pos1.x0, pos1.y0, pos1.width , 0.15]
    ax.set_position(pos)
    ax2 = seaborn.heatmap(normalized2, ax=ax2, cmap ="YlGnBu",
                 cbar=False,vmin = vmin, vmax = vmax,
                 cbar_kws={"orientation": "horizontal"})
    pos2 = [pos1.x0, pos1.y0 - 0.14, pos1.width , 0.13] 
    ax2.set_position(pos2)
    ax2.set_ylabel("Photorespiration", rotation = 0, fontsize = 16, color="k")
    ax2.yaxis.set_label_coords(-0.35, 0.5)
    ax2.axes.xaxis.set_visible(False)
    ax3 = seaborn.heatmap(normalized3, ax=ax3, cmap ="YlGnBu",
                 cbar=False,vmin = vmin, vmax = vmax,
                 cbar_kws={"orientation": "horizontal"})
    ax3.set_ylabel("TCA and\nOxidative\nPhosphorylation", rotation = 0, fontsize = 16, color="k")
    ax3.yaxis.set_label_coords(-0.35, 0.35)
    ax3.axes.xaxis.set_visible(False)
    pos2 = ax2.get_position()
    pos3 = [pos2.x0, pos2.y0 - 0.09, pos2.width , 0.08] 
    ax3.set_position(pos3)
    # ax4 = seaborn.heatmap(normalized4, ax=ax4, cmap ="YlGnBu"
    #              ,vmin = vmin, vmax = vmax, cbar_ax=cbar_ax,
    #              cbar_kws={"orientation": "horizontal"})
    # ax4.set_yticklabels(ax4.get_yticklabels(), rotation = 0)
    # ax4.set_ylabel("GS-GOGAT", rotation = 0, fontsize = 16, color="k")
    # ax4.yaxis.set_label_coords(-0.35, 0.3)
    # ax4.axes.xaxis.set_visible(False)
    # pos3 = ax3.get_position()
    # pos4 = [pos3.x0, pos3.y0 - 0.06, pos3.width , 0.05]
    # ax4.set_position(pos4)

    ax5 = seaborn.heatmap(normalized5, ax=ax5, cmap="YlGnBu"
                          , vmin=vmin, vmax=vmax, cbar_ax=cbar_ax,
                          cbar_kws={"orientation": "horizontal"})
    ax5.set_yticklabels(ax5.get_yticklabels(), rotation=0)
    ax5.set_ylabel("MEP\nPathway", rotation=0, fontsize=16, color="k")
    ax5.yaxis.set_label_coords(-0.35, 0.25)
    pos4 = ax3.get_position()
    pos5 = [pos4.x0, pos4.y0 - 0.06, pos4.width , 0.05]
    ax5.set_xlabel("Vc/Vo",  fontsize = 20, color="k")
    ax5.xaxis.set_label_coords(0.5, -0.3)
    ax5.set_position(pos5)
    pos5 = ax5.get_position()
    cbar_ax.set_position([pos5.x0, pos5.y0 - 0.08, pos5.width , 0.01])
    cbar_ax.title.set_text("Fold change in fluxes")

